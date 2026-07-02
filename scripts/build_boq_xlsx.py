"""
build_boq_xlsx.py - GIAI ĐOẠN 1: xuất file MỜI THẦU (BOQ trống giá).

Từ cau-hinh.json + 02-boq/<phong>.csv -> 03-baogia/moi-thau.xlsx:
  - Mỗi tầng 1 sheet, các phòng cùng tầng nằm liên tiếp trong sheet đó.
  - NCC điền 2 cột: ĐG VẬT LIỆU + ĐG NHÂN CÔNG (chào tách để so sánh thầu theo thành phần).
  - ĐƠN GIÁ = VL+NC (công thức); THÀNH TIỀN = KL*đơn giá (công thức).
  - KHÔNG chứa profit / giá nội bộ.

Dùng:  python scripts/build_boq_xlsx.py <project_dir>
"""
import os
import sys

from openpyxl import Workbook

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from lib_boq import (  # noqa: E402
    load_config, load_room_rows, group_rows, style_header_row,
    set_widths, roman, safe_save, place_image, load_image_map, minhhoa_col_width,
    MONEY, QTY, BOLD, GRP_FILL, TOTAL_FILL, BORDER,
)
from openpyxl.utils import get_column_letter  # noqa: E402

# Cột (1-based) — file mời thầu. Dùng hằng số để không lệch chỉ số khi sửa layout.
C_STT, C_KYHIEU, C_HANGMUC, C_QUYCACH, C_MINHHOA, C_DVT, C_DIENGIAI, \
    C_KL, C_TONGKL, C_DGVL, C_DGNC, C_DONGIA, C_TT1, C_TTDU, C_GHICHU = range(1, 16)
NCOL = C_GHICHU

HEADERS = [
    "STT", "KÝ HIỆU", "HẠNG MỤC / MÔ TẢ CÔNG VIỆC", "QUY CÁCH / THÔNG SỐ KỸ THUẬT",
    "MINH HỌA", "ĐVT", "DIỄN GIẢI (PHÂN TÍCH KL)", "KHỐI LƯỢNG", "TỔNG KHỐI LƯỢNG",
    "ĐG VẬT LIỆU (VND)", "ĐG NHÂN CÔNG (VND)", "ĐƠN GIÁ = VL+NC",
    "THÀNH TIỀN 1 PHÒNG", "THÀNH TIỀN ĐỦ PHÒNG", "GHI CHÚ / NGUỒN GỐC",
]
WIDTHS = [6, 10, 32, 26, 10, 7, 28, 10, 12, 13, 13, 14, 15, 16, 22]


def _L(col):
    return get_column_letter(col)


FLOOR_BY_ROOM = {
    "GT": "T.HẦM", "GARA": "T.HẦM", "WC0": "T.HẦM", "KHO-KT": "T.HẦM",
    "BEP": "T1", "LIVING": "T1", "POWDER": "T1",
    "PN1": "T2", "WC1": "T2", "THUVIEN": "T2",
    "PN2": "T3", "WC3": "T3", "MASTER": "T3", "WC4": "T3",
    "WORKSHOP": "TUM", "SANPHOI": "TUM",
}


def create_floor_sheet(wb, cfg, floor):
    ws = wb.create_sheet(floor[:31])
    ws["A1"] = f"DỰ ÁN: {cfg['du_an']}"
    ws["A2"] = f"HẠNG MỤC: {cfg.get('hang_muc', '')}"
    ws["A3"] = f"TẦNG: {floor}"
    ws["A1"].font = ws["A2"].font = ws["A3"].font = BOLD
    header_row = 4
    for j, h in enumerate(HEADERS, 1):
        ws.cell(header_row, j, h)
    style_header_row(ws, header_row, NCOL)
    set_widths(ws, WIDTHS)
    ws.freeze_panes = "A5"
    return ws, header_row + 1


def append_room(ws, cfg, phong, rows, row, images=None):
    images = images or {}
    sl = phong["so_luong"]
    ws.cell(row, C_STT, "A")
    ws.cell(row, C_HANGMUC, f"Phòng {phong['ma']} ({phong['ten']})")
    ws.cell(row, C_QUYCACH, f"{sl} phòng")
    ws.cell(row, C_TONGKL, sl)
    for c in range(1, NCOL + 1):
        ws.cell(row, c).font = BOLD
        ws.cell(row, c).fill = TOTAL_FILL
    row += 1

    stt = 0
    first_item = row
    last_item = row
    for gi, (_nm, group) in enumerate(group_rows(rows, cfg["scope"]), 1):
        ws.cell(row, C_STT, roman(gi))
        ws.cell(row, C_HANGMUC, group["ten"]).font = BOLD
        for c in range(1, NCOL + 1):
            ws.cell(row, c).fill = GRP_FILL
        row += 1

        for it in group["items"]:
            stt += 1
            kl = it["_kl"]
            ws.cell(row, C_STT, stt)
            ws.cell(row, C_KYHIEU, it.get("ky_hieu"))
            ws.cell(row, C_HANGMUC, it.get("hang_muc"))
            ws.cell(row, C_QUYCACH, it.get("quy_cach"))
            ws.cell(row, C_DVT, it.get("don_vi"))
            ws.cell(row, C_DIENGIAI, it.get("dien_giai"))
            if kl is not None:
                ws.cell(row, C_KL, kl).number_format = QTY
                ws.cell(row, C_TONGKL, f"={_L(C_KL)}{row}*{sl}").number_format = QTY
            # NCC điền VL + NC; đơn giá = VL+NC; thành tiền = KL * đơn giá (đều là công thức).
            ws.cell(row, C_DGVL).number_format = MONEY
            ws.cell(row, C_DGNC).number_format = MONEY
            ws.cell(row, C_DONGIA, f"={_L(C_DGVL)}{row}+{_L(C_DGNC)}{row}").number_format = MONEY
            ws.cell(row, C_TT1, f"={_L(C_KL)}{row}*{_L(C_DONGIA)}{row}").number_format = MONEY
            ws.cell(row, C_TTDU, f"={_L(C_TONGKL)}{row}*{_L(C_DONGIA)}{row}").number_format = MONEY
            ws.cell(row, C_GHICHU, it.get("ghi_chu"))
            for c in range(1, NCOL + 1):
                ws.cell(row, c).border = BORDER
            img = images.get((it.get("hang_muc") or "").strip())
            if img:
                place_image(ws, row, img, col_idx=C_MINHHOA)
            last_item = row
            row += 1

    ws.cell(row, C_HANGMUC, "TỔNG PHÒNG TRƯỚC VAT").font = BOLD
    ws.cell(row, C_TT1, f"=SUM({_L(C_TT1)}{first_item}:{_L(C_TT1)}{last_item})").number_format = MONEY
    ws.cell(row, C_TTDU, f"=SUM({_L(C_TTDU)}{first_item}:{_L(C_TTDU)}{last_item})").number_format = MONEY
    ws.cell(row, C_TT1).font = ws.cell(row, C_TTDU).font = BOLD
    total_row = row
    return row + 2, total_row, stt


def build(project_dir):
    project_dir = str(project_dir).rstrip("\\/")
    cfg = load_config(project_dir)
    wb = Workbook()
    wb.remove(wb.active)
    out_dir = os.path.join(project_dir, "03-baogia")
    os.makedirs(out_dir, exist_ok=True)

    floors = []
    floor_rooms = {}
    for phong in cfg["phong"]:
        floor = FLOOR_BY_ROOM.get(phong["ma"], "KHÁC")
        if floor not in floor_rooms:
            floors.append(floor)
            floor_rooms[floor] = []
        floor_rooms[floor].append(phong)

    missing = []
    total_no_price = 0
    n_img = 0
    for floor in floors:
        ws, row = create_floor_sheet(wb, cfg, floor)
        total_rows = []
        has_data = False
        floor_has_img = False
        for phong in floor_rooms[floor]:
            rows = load_room_rows(project_dir, phong["ma"])
            if rows is None:
                missing.append(phong["ma"])
                continue
            total_no_price += sum(1 for r in rows if r.get("_gia_eff") is None)
            images = load_image_map(project_dir, phong["ma"])
            row, total_row, count = append_room(ws, cfg, phong, rows, row, images)
            total_rows.append(total_row)
            has_data = True
            if images:
                floor_has_img = True
                n_img += sum(1 for v in images.values() if v)
            print(f"  {floor}/{phong['ma']}: {count} hạng mục"
                  + (f" (+{len(images)} ảnh)" if images else ""))
        if floor_has_img:
            ws.column_dimensions[_L(C_MINHHOA)].width = minhhoa_col_width()

        if has_data and total_rows:
            ws.cell(row, C_HANGMUC, f"TỔNG CỘNG {floor} TRƯỚC VAT").font = BOLD
            ws.cell(row, C_TT1, "=" + "+".join(f"{_L(C_TT1)}{x}" for x in total_rows)).number_format = MONEY
            ws.cell(row, C_TTDU, "=" + "+".join(f"{_L(C_TTDU)}{x}" for x in total_rows)).number_format = MONEY
            ws.cell(row, C_TT1).font = ws.cell(row, C_TTDU).font = BOLD
            for c in range(1, NCOL + 1):
                ws.cell(row, c).fill = TOTAL_FILL
        else:
            del wb[ws.title]

    if not wb.sheetnames:
        raise ValueError(f"Không có file 02-boq/*.csv. Thiếu: {missing}")
    out = safe_save(wb, os.path.join(out_dir, "moi-thau.xlsx"))
    print(f"\n[GĐ1] File mời thầu (trống giá) -> {out}")
    if n_img:
        print(f"  Đã chèn {n_img} ảnh minh họa vào cột MINH HỌA.")
    if missing:
        print(f"  (Chưa có BOQ cho phòng: {missing})")
    print("  Gửi file này cho nhà thầu phụ điền ĐG VẬT LIỆU + ĐG NHÂN CÔNG. KHÔNG chứa profit.")
    return out, {"n_no_price": total_no_price, "n_img": n_img}


def main():
    if len(sys.argv) < 2:
        sys.exit("Dùng: python scripts/build_boq_xlsx.py <project_dir>")
    try:
        build(sys.argv[1])
    except Exception as e:
        sys.exit(str(e))


if __name__ == "__main__":
    main()
