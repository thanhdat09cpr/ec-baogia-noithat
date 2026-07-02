"""
build_baogia_xlsx.py - GIAI ĐOẠN 2: xuất BÁO GIÁ NỘI BỘ (có profit).

Từ cau-hinh.json + 02-boq/<phong>.csv (đã có giá NCC) -> 03-baogia/bao-gia-noi-bo.xlsx:
  - Mỗi tầng 1 sheet, các phòng cùng tầng nằm liên tiếp trong sheet đó.
  - giá_NCC = don_gia_ncc (trọn gói) hoặc don_gia_vl + don_gia_nc (chào tách).
  - Đơn giá báo giá = giá_NCC * (1 + profit%). File nội bộ chỉ hiện 1 cột đơn giá (profit ẩn).
  - Sheet TH tổng hợp theo từng phòng + Preliminaries + VAT.

Dùng:  python scripts/build_baogia_xlsx.py <project_dir> [--profit 10]
"""
import argparse
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from lib_boq import (  # noqa: E402
    load_config, load_room_rows, group_rows, style_header_row,
    set_widths, roman, safe_save, place_image, load_image_map, minhhoa_col_width,
    MONEY, QTY, BOLD, GRP_FILL, TOTAL_FILL, BORDER,
)
from openpyxl.utils import get_column_letter  # noqa: E402

# Cột (1-based) — báo giá nội bộ. 1 cột đơn giá (profit ẩn); có ký hiệu + diễn giải.
C_STT, C_KYHIEU, C_HANGMUC, C_QUYCACH, C_MINHHOA, C_DVT, C_DIENGIAI, \
    C_KL, C_TONGKL, C_DONGIA, C_TT1, C_TTDU, C_GHICHU = range(1, 14)
NCOL = C_GHICHU

HEADERS = [
    "STT", "KÝ HIỆU", "HẠNG MỤC / MÔ TẢ CÔNG VIỆC", "QUY CÁCH / THÔNG SỐ KỸ THUẬT",
    "MINH HỌA", "ĐVT", "DIỄN GIẢI (PHÂN TÍCH KL)", "KHỐI LƯỢNG", "TỔNG KHỐI LƯỢNG",
    "ĐƠN GIÁ BÁO GIÁ (VND)", "THÀNH TIỀN 1 PHÒNG", "THÀNH TIỀN ĐỦ PHÒNG",
    "GHI CHÚ / NGUỒN GỐC",
]
WIDTHS = [6, 10, 32, 26, 10, 7, 28, 10, 12, 16, 16, 17, 22]


def _L(col):
    return get_column_letter(col)


FLOOR_BY_ROOM = {
    "GT": "T.HẦM", "GARA": "T.HẦM", "WC0": "T.HẦM", "KHO-KT": "T.HẦM",
    "BEP": "T1", "LIVING": "T1", "POWDER": "T1",
    "PN1": "T2", "WC1": "T2", "THUVIEN": "T2",
    "PN2": "T3", "WC3": "T3", "MASTER": "T3", "WC4": "T3",
    "WORKSHOP": "TUM", "SANPHOI": "TUM",
}


def font_big():
    return Font(bold=True, size=12)


def create_floor_sheet(wb, cfg, floor):
    ws = wb.create_sheet(floor[:31])
    ws["A1"] = f"DỰ ÁN: {cfg['du_an']}"
    ws["A2"] = f"HẠNG MỤC: {cfg.get('hang_muc', '')}"
    ws["A3"] = f"TẦNG: {floor}"
    ws["A1"].font = ws["A2"].font = ws["A3"].font = BOLD
    header_row = 4
    for j, h in enumerate(HEADERS, 1):
        ws.cell(header_row, j, h)
    style_header_row(ws, header_row, NCOL)
    set_widths(ws, WIDTHS)
    ws.freeze_panes = "A5"
    return ws, header_row + 1


def append_room(ws, cfg, phong, rows, row, profit_default, images=None):
    images = images or {}
    sl = phong["so_luong"]
    ws.cell(row, C_STT, "A")
    ws.cell(row, C_HANGMUC, f"Phòng {phong['ma']} ({phong['ten']})")
    ws.cell(row, C_QUYCACH, f"{sl} phòng")
    ws.cell(row, C_TONGKL, sl)
    room_header_row = row
    for c in range(1, NCOL + 1):
        ws.cell(row, c).font = BOLD
        ws.cell(row, c).fill = TOTAL_FILL
    row += 1

    stt = 0
    first_item = row
    n_no_price = 0
    for gi, (_nm, group) in enumerate(group_rows(rows, cfg["scope"]), 1):
        ws.cell(row, C_STT, roman(gi))
        ws.cell(row, C_HANGMUC, group["ten"]).font = BOLD
        for c in range(1, NCOL + 1):
            ws.cell(row, c).fill = GRP_FILL
        row += 1

        for it in group["items"]:
            stt += 1
            kl = it["_kl"]
            ncc = it["_gia_eff"]   # giá NCC hiệu dụng: trọn gói hoặc VL+NC
            profit = it["_profit_override"]
            profit = profit if profit is not None else profit_default
            gia = ncc * (1 + profit / 100.0) if ncc is not None else None

            ws.cell(row, C_STT, stt)
            ws.cell(row, C_KYHIEU, it.get("ky_hieu"))
            ws.cell(row, C_HANGMUC, it.get("hang_muc"))
            ws.cell(row, C_QUYCACH, it.get("quy_cach"))
            ws.cell(row, C_DVT, it.get("don_vi"))
            ws.cell(row, C_DIENGIAI, it.get("dien_giai"))
            if kl is not None:
                ws.cell(row, C_KL, kl).number_format = QTY
                ws.cell(row, C_TONGKL, f"={_L(C_KL)}{row}*{sl}").number_format = QTY
            if gia is not None:
                ws.cell(row, C_DONGIA, round(gia)).number_format = MONEY
            else:
                n_no_price += 1
            ws.cell(row, C_TT1, f"={_L(C_KL)}{row}*{_L(C_DONGIA)}{row}").number_format = MONEY
            ws.cell(row, C_TTDU, f"={_L(C_TONGKL)}{row}*{_L(C_DONGIA)}{row}").number_format = MONEY
            note = it.get("ghi_chu") or ""
            if ncc is None:
                note = (note + " | CHƯA CÓ GIÁ NCC").strip(" |")
            ws.cell(row, C_GHICHU, note)
            for c in range(1, NCOL + 1):
                ws.cell(row, c).border = BORDER
            img = images.get((it.get("hang_muc") or "").strip())
            if img:
                place_image(ws, row, img, col_idx=C_MINHHOA)
            row += 1

    last_item = row - 1
    ws.cell(row, C_HANGMUC, "TỔNG PHÒNG TRƯỚC VAT").font = BOLD
    ws.cell(row, C_TT1, f"=SUM({_L(C_TT1)}{first_item}:{_L(C_TT1)}{last_item})").number_format = MONEY
    ws.cell(row, C_TTDU, f"=SUM({_L(C_TTDU)}{first_item}:{_L(C_TTDU)}{last_item})").number_format = MONEY
    ws.cell(row, C_TT1).font = ws.cell(row, C_TTDU).font = BOLD
    ws.cell(room_header_row, C_TT1, f"={_L(C_TT1)}{row}").number_format = MONEY
    ws.cell(room_header_row, C_TTDU, f"={_L(C_TTDU)}{row}").number_format = MONEY
    return row + 2, row, stt, n_no_price


def build_th(wb, cfg, room_refs):
    th = wb.create_sheet("TH", 0)
    th["A1"] = "BẢNG TỔNG HỢP BÁO GIÁ"
    th["A1"].font = font_big()
    th["A2"] = f"DỰ ÁN: {cfg['du_an']}"
    th["A3"] = f"ĐỊA ĐIỂM: {cfg.get('dia_diem', '')}"
    th["A4"] = f"HẠNG MỤC: {cfg.get('hang_muc', '')}"
    header_row = 6
    headers = ["STT", "PHÒNG / KHOẢN MỤC", "SỐ LƯỢNG",
               "GIÁ TRỊ 1 PHÒNG (CHƯA VAT)", "TỔNG GIÁ TRỊ (CHƯA VAT)", "GHI CHÚ"]
    for j, h in enumerate(headers, 1):
        th.cell(header_row, j, h)
    style_header_row(th, header_row, 6)

    row = header_row + 1
    first = row
    for i, (sheet, total_row, phong) in enumerate(room_refs, 1):
        th.cell(row, 1, i)
        th.cell(row, 2, phong["ten"])
        th.cell(row, 3, phong["so_luong"])
        th.cell(row, 4, f"='{sheet}'!{_L(C_TT1)}{total_row}").number_format = MONEY
        th.cell(row, 5, f"='{sheet}'!{_L(C_TTDU)}{total_row}").number_format = MONEY
        for c in range(1, 7):
            th.cell(row, c).border = BORDER
        row += 1
    last = row - 1

    th.cell(row, 2, "B. CỘNG CHI PHÍ THI CÔNG (các phòng)").font = BOLD
    th.cell(row, 5, f"=SUM(E{first}:E{last})").number_format = MONEY
    th.cell(row, 5).font = BOLD
    thi_cong_row = row
    row += 1

    th.cell(row, 2, f"A. {cfg.get('preliminaries_ten')}")
    th.cell(row, 5, cfg.get("preliminaries_lumpsum") or 0).number_format = MONEY
    th.cell(row, 6, "Nhập tay trọn gói")
    prelim_row = row
    row += 1

    th.cell(row, 2, "TỔNG GIÁ TRỊ TRƯỚC VAT").font = BOLD
    th.cell(row, 5, f"=E{thi_cong_row}+E{prelim_row}").number_format = MONEY
    th.cell(row, 5).font = BOLD
    before_vat_row = row
    row += 1

    vat = cfg.get("vat_percent", 8)
    th.cell(row, 2, f"THUẾ VAT {vat}%")
    th.cell(row, 4, vat / 100.0).number_format = "0%"
    th.cell(row, 5, f"=E{before_vat_row}*D{row}").number_format = MONEY
    vat_row = row
    row += 1

    th.cell(row, 2, "TỔNG GIÁ TRỊ BAO GỒM VAT").font = font_big()
    th.cell(row, 5, f"=E{before_vat_row}+E{vat_row}").number_format = MONEY
    th.cell(row, 5).font = font_big()
    for c in range(1, 7):
        th.cell(row, c).fill = TOTAL_FILL
    set_widths(th, [5, 42, 9, 22, 22, 24])


def build(project_dir, profit=None):
    project_dir = str(project_dir).rstrip("\\/")
    cfg = load_config(project_dir)
    profit = profit if profit is not None else cfg["profit_percent"]
    print(f"Profit mặc định: {profit}% | VAT: {cfg['vat_percent']}% | "
          f"Preliminaries: {cfg.get('preliminaries_lumpsum'):,}")

    floors = []
    floor_rooms = {}
    for phong in cfg["phong"]:
        floor = FLOOR_BY_ROOM.get(phong["ma"], "KHÁC")
        if floor not in floor_rooms:
            floors.append(floor)
            floor_rooms[floor] = []
        floor_rooms[floor].append(phong)

    wb = Workbook()
    wb.remove(wb.active)
    room_refs = []
    total_no_price = 0
    n_img = 0
    for floor in floors:
        ws, row = create_floor_sheet(wb, cfg, floor)
        total_rows = []
        has_data = False
        floor_has_img = False
        for phong in floor_rooms[floor]:
            rows = load_room_rows(project_dir, phong["ma"])
            if rows is None:
                print(f"  (bỏ qua {phong['ma']} - chưa có 02-boq/{phong['ma']}.csv)")
                continue
            images = load_image_map(project_dir, phong["ma"])
            row, total_row, count, no_price = append_room(
                ws, cfg, phong, rows, row, profit, images)
            total_rows.append(total_row)
            room_refs.append((ws.title, total_row, phong))
            total_no_price += no_price
            has_data = True
            if images:
                floor_has_img = True
                n_img += sum(1 for v in images.values() if v)
            suffix = f" ({no_price} dòng chưa có giá)" if no_price else ""
            print(f"  {floor}/{phong['ma']}: {count} hạng mục{suffix}"
                  + (f" (+{len(images)} ảnh)" if images else ""))
        if floor_has_img:
            ws.column_dimensions[_L(C_MINHHOA)].width = minhhoa_col_width()

        if has_data and total_rows:
            ws.cell(row, C_HANGMUC, f"TỔNG CỘNG {floor} TRƯỚC VAT").font = BOLD
            ws.cell(row, C_TT1, "=" + "+".join(f"{_L(C_TT1)}{x}" for x in total_rows)).number_format = MONEY
            ws.cell(row, C_TTDU, "=" + "+".join(f"{_L(C_TTDU)}{x}" for x in total_rows)).number_format = MONEY
            ws.cell(row, C_TT1).font = ws.cell(row, C_TTDU).font = BOLD
            for c in range(1, NCOL + 1):
                ws.cell(row, c).fill = TOTAL_FILL
        else:
            del wb[ws.title]

    if not room_refs:
        raise ValueError("Không có file 02-boq/*.csv.")
    build_th(wb, cfg, room_refs)
    out_dir = os.path.join(project_dir, "03-baogia")
    os.makedirs(out_dir, exist_ok=True)
    out = safe_save(wb, os.path.join(out_dir, "bao-gia-noi-bo.xlsx"))
    print(f"\n[GĐ2] Báo giá nội bộ (có profit {profit}%) -> {out}")
    if n_img:
        print(f"  Đã chèn {n_img} ảnh minh họa vào cột MINH HỌA.")
    if total_no_price:
        print(f"  Cảnh báo: {total_no_price} dòng CHƯA CÓ đơn giá NCC.")
    return out, {"n_no_price": total_no_price, "n_img": n_img}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("project")
    ap.add_argument("--profit", type=float, default=None, help="ghi đè profit%% chung")
    args = ap.parse_args()
    try:
        build(args.project, args.profit)
    except Exception as e:
        sys.exit(str(e))


if __name__ == "__main__":
    main()
