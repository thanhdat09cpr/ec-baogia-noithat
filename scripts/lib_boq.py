"""
lib_boq.py — Tien ich dung chung cho build_boq_xlsx.py va build_baogia_xlsx.py.

Doc cau hinh du an + cac file BOQ tung phong, va cung cap style Excel.
"""
import os, re, csv, json
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

CSV_COLS = ["nhom_ma", "nhom_ten", "hang_muc", "quy_cach", "don_vi",
            "kl_1phong", "don_gia_ncc", "do_tin_cay", "ghi_chu"]

MONEY = "#,##0"
QTY = "#,##0.###"
BOLD = Font(bold=True)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF")
GRP_FILL = PatternFill("solid", fgColor="DDEBF7")
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _looks_like_float_artifact(frac):
    if len(frac) <= 2:
        return False
    stripped = frac.rstrip("0")
    if not stripped:
        return True
    if set(frac) == {"9"}:
        return True
    return frac.startswith("000000") or frac.startswith("999999")


def to_number(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"\s+", "", str(v))
    s = re.sub(r"[^\d.,-]", "", s)
    if not s:
        return None
    neg = s.startswith("-")
    s = s.replace("-", "")
    if re.fullmatch(r"\d{1,3}([.,]\d{3})+", s):
        s = re.sub(r"[.,]", "", s)
    elif s.count(".") + s.count(",") == 1:
        sep = "." if "." in s else ","
        whole, frac = s.rsplit(sep, 1)
        if not whole.isdigit() or not frac.isdigit():
            return None
        if len(frac) <= 2 or (sep == "." and len(whole) > 3 and _looks_like_float_artifact(frac)):
            s = whole + "." + frac
        else:
            return None
    elif not s.isdigit():
        return None
    if neg:
        s = "-" + s
    try:
        return float(s)
    except ValueError:
        return None


def to_quantity(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"\s+", "", str(v))
    s = re.sub(r"[^\d.,-]", "", s)
    if not s:
        return None
    neg = s.startswith("-")
    s = s.replace("-", "")
    if s.count(".") + s.count(",") == 1:
        sep = "." if "." in s else ","
        whole, frac = s.rsplit(sep, 1)
        if whole.isdigit() and frac.isdigit() and len(frac) <= 3:
            s = whole + "." + frac
        elif re.fullmatch(r"\d{1,3}([.,]\d{3})+", s):
            s = re.sub(r"[.,]", "", s)
        else:
            return None
    elif re.fullmatch(r"\d{1,3}([.,]\d{3})+", s):
        s = re.sub(r"[.,]", "", s)
    elif not s.isdigit():
        return None
    if neg:
        s = "-" + s
    try:
        return float(s)
    except ValueError:
        return None


def load_config(project_dir):
    p = os.path.join(project_dir, "cau-hinh.json")
    if not os.path.exists(p):
        raise FileNotFoundError(f"Thieu {p}. Copy tu templates/cau-hinh.mau.json va chinh.")
    with open(p, encoding="utf-8-sig") as f:
        cfg = json.load(f)
    cfg.setdefault("profit_percent", 10)
    cfg.setdefault("vat_percent", 8)
    cfg.setdefault("preliminaries_lumpsum", 0)
    cfg.setdefault("preliminaries_ten", "CÔNG TÁC CHUẨN BỊ (PRELIMINARIES)")
    cfg.setdefault("scope", ["I.1", "I.2", "I.3", "I.4", "I.5"])
    return cfg


def load_room_rows(project_dir, ma):
    """Doc 02-boq/<ma>.csv -> list dict (giu thu tu). Tra ve [] neu khong co file."""
    p = os.path.join(project_dir, "02-boq", f"{ma}.csv")
    if not os.path.exists(p):
        return None
    rows = []
    with open(p, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            r = {k: (r.get(k) or "").strip() for k in r}
            r["_kl"] = to_quantity(r.get("kl_1phong"))
            r["_gia_ncc"] = to_number(r.get("don_gia_ncc"))
            r["_profit_override"] = to_number(r.get("profit_override"))
            rows.append(r)
    return rows


def group_rows(rows, scope):
    """Gom theo nhom_ma, theo thu tu scope roi theo thu tu xuat hien."""
    order, groups = [], {}
    for r in rows:
        nm = r.get("nhom_ma") or "?"
        if nm not in groups:
            groups[nm] = {"ten": r.get("nhom_ten") or nm, "items": []}
            order.append(nm)
        groups[nm]["items"].append(r)
    def keyf(nm):
        return (scope.index(nm) if nm in scope else 999, order.index(nm))
    return [(nm, groups[nm]) for nm in sorted(order, key=keyf)]


def style_header_row(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row, c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def set_widths(ws, widths):
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def safe_save(wb, path):
    """Luu workbook; neu file dang mo (PermissionError) -> luu ban moi co timestamp."""
    import time
    try:
        wb.save(path)
        return path
    except PermissionError:
        base, ext = os.path.splitext(path)
        alt = f"{base}_{time.strftime('%H%M%S')}{ext}"
        wb.save(alt)
        print(f"  ⚠ '{os.path.basename(path)}' đang mở trong Excel — đã lưu '{os.path.basename(alt)}'.")
        return alt


def roman(n):
    vals = [(10, "X"), (9, "IX"), (5, "V"), (4, "IV"), (1, "I")]
    out = ""
    for v, s in vals:
        while n >= v:
            out += s; n -= v
    return out


# ---- Chen anh vao cot MINH HOA (crop vua du, fit trong o, khong vo bo cuc) ----
MINHHOA_BOX = (122, 112)          # (rong, cao) px toi da cua o anh
MINHHOA_COL_CHARS = 18.5          # be rong cot D (~130px) du chua anh


def minhhoa_col_width():
    return MINHHOA_COL_CHARS


def _fit_image(path, box):
    """Mo anh, cat vien trang, resize vua trong box (giu ti le). Tra (BytesIO,w,h)."""
    import io
    from PIL import Image, ImageChops
    im = Image.open(path).convert("RGB")
    # cat vien don sac (an toan neu anh chua duoc crop san)
    bg = Image.new("RGB", im.size, (255, 255, 255))
    diff = ImageChops.difference(im, bg).point(lambda p: 255 if p > 12 else 0)
    bb = diff.getbbox()
    if bb:
        l, t, r, b = bb
        pad = 4
        im = im.crop((max(0, l - pad), max(0, t - pad),
                      min(im.width, r + pad), min(im.height, b + pad)))
    bw, bh = box
    scale = min(bw / im.width, bh / im.height)
    w, h = max(1, int(im.width * scale)), max(1, int(im.height * scale))
    im = im.resize((w, h), Image.LANCZOS)
    buf = io.BytesIO(); im.save(buf, format="PNG"); buf.seek(0)
    return buf, w, h


def place_image(ws, row, img_path, col_idx=4, box=MINHHOA_BOX):
    """Chen anh vao o (col_idx,row), can giua, va NOI RONG dong vua du anh.
    Tra True neu chen duoc. Neu file khong ton tai -> False (bo qua)."""
    if not img_path or not os.path.exists(img_path):
        return False
    try:
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU
        buf, w, h = _fit_image(img_path, box)
        img = XLImage(buf)
        img.width, img.height = w, h
        # can giua trong o (cot ~130px, box cao)
        offx = max(0, (130 - w) // 2)
        offy = 3
        marker = AnchorMarker(col=col_idx - 1, colOff=pixels_to_EMU(offx),
                              row=row - 1, rowOff=pixels_to_EMU(offy))
        img.anchor = OneCellAnchor(_from=marker,
                                   ext=XDRPositiveSize2D(pixels_to_EMU(w), pixels_to_EMU(h)))
        ws.add_image(img)
        need = h * 0.75 + 6          # px -> point + padding
        cur = ws.row_dimensions[row].height or 15
        if need > cur:
            ws.row_dimensions[row].height = need
        return True
    except Exception as e:
        print(f"    [anh] bo qua {os.path.basename(img_path)}: {e}")
        return False


def load_image_map(project_dir, room):
    """Doc 02-boq/<room>.anh.csv -> {hang_muc: abs_img_path}. {} neu khong co."""
    p = os.path.join(project_dir, "02-boq", room + ".anh.csv")
    if not os.path.exists(p):
        return {}
    base = os.path.join(project_dir, "01-extract")
    m = {}
    with open(p, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            ip = (r.get("img_path") or "").strip()
            hm = (r.get("hang_muc") or "").strip()
            if not ip or not hm:
                continue
            # cho phep ghi duong dan tuong doi (spec-img/..) hoac ten file
            cand = ip if os.path.isabs(ip) else os.path.join(base, ip)
            if not os.path.exists(cand):
                cand2 = os.path.join(base, "spec-img", os.path.basename(ip))
                cand = cand2 if os.path.exists(cand2) else cand
            m[hm] = cand
    return m
