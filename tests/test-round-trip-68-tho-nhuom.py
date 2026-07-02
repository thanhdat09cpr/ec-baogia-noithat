import os
import shutil
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from build_baogia_xlsx import build as build_baogia  # noqa: E402
from build_boq_xlsx import build as build_boq  # noqa: E402


PROJECT = ROOT / "projects" / "68-Tho-Nhuom"
REFERENCE_TOTAL = 1_302_963_612.5
REFERENCE_WITH_PROFIT_TOTAL = 1_433_259_978.75


def _make_project(csv_name):
    tmp = tempfile.TemporaryDirectory()
    project = Path(tmp.name) / "68-Tho-Nhuom"
    (project / "02-boq").mkdir(parents=True)
    shutil.copy(PROJECT / "cau-hinh.json", project / "cau-hinh.json")
    shutil.copy(PROJECT / "02-boq" / csv_name, project / "02-boq" / "GR1.csv")
    return tmp, project


def _item_total(path):
    wb = load_workbook(path, data_only=False)
    ws = wb["KHÁC"]
    total = 0
    for row in range(1, ws.max_row + 1):
        if isinstance(ws.cell(row, 1).value, int):
            quantity = ws.cell(row, 6).value or 0
            room_count = 5
            price = ws.cell(row, 8).value or 0
            total += quantity * room_count * price
    return total


class RoundTrip68ThoNhuomTest(unittest.TestCase):
    def test_moi_thau_build_keeps_price_columns_blank(self):
        tmp, project = _make_project("GR1.csv")
        with tmp:
            out_path, stats = build_boq(project)
            self.assertEqual(stats["n_no_price"], 0)
            self.assertEqual(stats["n_img"], 0)

            wb = load_workbook(out_path, data_only=False)
            ws = wb["KHÁC"]
            prices = [
                ws.cell(row, 8).value
                for row in range(1, ws.max_row + 1)
                if isinstance(ws.cell(row, 1).value, int)
            ]
            self.assertTrue(prices)
            self.assertTrue(all(value is None for value in prices))

    def test_bao_gia_round_trip_reference_total_without_profit(self):
        tmp, project = _make_project("GR1.thuc-tham-chieu.csv")
        with tmp:
            out_path, stats = build_baogia(project, profit=0)
            self.assertEqual(stats["n_no_price"], 0)
            self.assertEqual(stats["n_img"], 0)
            self.assertEqual(_item_total(out_path), REFERENCE_TOTAL)

    def test_bao_gia_round_trip_reference_total_with_default_profit(self):
        tmp, project = _make_project("GR1.thuc-tham-chieu.csv")
        with tmp:
            out_path, stats = build_baogia(project)
            self.assertEqual(stats["n_no_price"], 0)
            self.assertEqual(stats["n_img"], 0)
            self.assertEqual(_item_total(out_path), REFERENCE_WITH_PROFIT_TOTAL)


if __name__ == "__main__":
    unittest.main()
